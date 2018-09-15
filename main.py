import copy
import os

import  openpyxl
from typing import Container

TEMPLATE_DIRECTORY = r'F:\LabData\Lab\ISO 17025\Control Chart\Templates\fastmake ccs'
CHART_DIRECTORY = r'F:\LabData\Lab\ISO 17025\Control Chart\Templates\fastmake ccs\chart copies'
TARGET_DIRECTORY = r'F:\LabData\Lab\ISO 17025\Control Chart\Templates\fastmake ccs\chart copies\output'

class Chart:
    DEFAULT_INPUT_ADDRESSES = ('A2', 'B2', 'C2', 'D2',)
    OLD_STDEV_ADDRESS = 'AC2'
    ETS_RESULT_ADDRESS = 'AB2'

    def __init__(self, file_path, name=None):
        self.file_path = file_path
        self.name = name
        self.wb = None
        self.active_worksheet = None

    def __str__(self):
        re.search(r'[\|/]+([\w]+).x[\w]{2,3}', self.file_path)[0]

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(self.file_path)

    def set_active_worksheet(self, target):
        self.active_worksheet = wb.sheets[target]

    def get_cell_copy(self, target) -> openpyxl.cell.cell.Cell:
        return copy.deepcopy(self.active_worksheet[f'{target}'])

    def set_cell(self, target, value):
        self.active_worksheet[f'{target}'] = value


class ChartGatherer:
    def __init__(self, directory, chart_format=Chart):
        self.directory = directory
        self.gathered_charts: List[Chart] = []
        self.chart_format = chart_format

    def gather_charts(self, extension='xlsx'):
        for dirpath, dirnames, filenames in os.walk(self.directory):
            filenames = [self.chart_format(os.path.join(dirpath, file)) for file in filenames if file.endswith(extension)]
            self.gathered_charts.extend(filenames)

class Template:
    def __init__(self, type):
        self.type = type


if __name__ == '__main__':
    chart_source = input('What is the Chart source? (input directory or file or blank for default chart directory)'
                         '\n'
                         '--> ')

    chart_source = chart_source if chart_source else CHART_DIRECTORY
    chart_gatherer = ChartGatherer(CHART_DIRECTORY)
    chart_gatherer.gather_charts()


    lcs_id = None
    continue_status = None
    for chart in chart_gatherer.gathered_charts:
        chart.load_workbook()
        if continue_status.lower() != 'yy':
            continue_status = input(f'Modify {chart}? (Y/N/NN for no to all/YY for yes to all)')

        if continue_status.lower() == 'n':
            continue
        elif continue_status.lower() == 'nn':
            break

        chart.set_active_worksheet(target=0)
        lcs_id_temp = input('What is the LCS ID? (input nothing to re-use last input)'
                       '\n'
                       '--> ')

        lcs_id = lcs_id_temp if lcs_id_temp else lcs_id

        chart.active_worksheet.title = lcs_id


        ets_result = input('What is the ETS value? (input nothing to re-use last input)'
                       '\n'
                       '--> ')

        chart.set_cell(Chart.ETS_RESULT_ADDRESS, ets_result)
        chart.set_cell(Chart.OLD_STDEV_ADDRESS, chart.get_cell_copy(Chart.OLD_STDEV_ADDRESS).value)

        input('Input starting data points formatted as mm/dd/yy, hhmm, initials, measured value:')
        while True:
            user_in = input('\n--> ').split()
            if not user_in:
                break

            for i, val in enumerate(user_in, int=0):
                chart.set_cell(Chart.DEFAULT_INPUT_ADDRESSES[i], val)

        chart.wb.save(chart + '.xlsx', as_template=False)